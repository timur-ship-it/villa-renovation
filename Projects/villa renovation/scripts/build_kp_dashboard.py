from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import quote
import html
import re

from openpyxl import load_workbook

EXCEL_PATH = Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/Scope_Manual_Input_from_TZ.xlsx')
OUTPUT_HTML = Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП сравнение.html')
OUTPUT_PDF_NAME = 'kp-sravnenie-live.pdf'
VILLA_AREA_M2 = 700.0
DROPBOX_SHARED_FOLDER_URL = 'https://www.dropbox.com/scl/fo/oqjk2cj2hpywhnofiy6gj/AFtZqpiQ_rO-mR4fdm9T-Nk?rlkey=vk3kaa8xd7zow6xu99uzfhl96&st=6kxd3zjs&dl=0'

# Подтверждено HTTP-проверкой (title != "No Access").
DROPBOX_KP_LINKS = {
    'D5 (AED)': 'https://www.dropbox.com/scl/fo/oifx7xmuxncq6h19w9e46/AEVj5ALoii0_0Tz3afSdX98?rlkey=iyt8rzh66tfhaje4isijnra1s&st=9xarcbg4&dl=0',
    'Smart (AED)': 'https://www.dropbox.com/scl/fo/c9v17ay7n2hqrd4ev4nlu/APFkV8Sqj9vBArovlR7rnhw?rlkey=oe5dyg2joatdnjaimt47xn24m&st=h442ofup&dl=0',
    'Антонович (AED)': 'https://www.dropbox.com/scl/fo/hlq4nu0ysxqxgjcuekdfs/ACUctLKZFFRiF0XJZSbXEwU?rlkey=0fjabdq0qutavere9loda88eo&st=3y6ic03h&dl=0',
}

DROPBOX_PORTFOLIO_LINKS = {
    'D5 (AED)': 'https://www.dropbox.com/scl/fo/oqjk2cj2hpywhnofiy6gj/AOsp58nbR_ROM0cbGmuPUqg/D5/%D0%9F%D0%BE%D1%80%D1%82%D1%84%D0%BE%D0%BB%D0%B8%D0%BE%20D5?dl=0&rlkey=vk3kaa8xd7zow6xu99uzfhl96&st=6kxd3zjs',
}

# Confirmed from source quote files (excl. VAT where explicitly present)
QUOTE_TOTALS = {
    'TX7 (AED)': 449300.0,
    'Фирсов (AED)': 640000.0,
    'D5 (AED)': 976000.0,
    'K4 (AED)': 1424745.05,
    'Smart (AED)': 755652.0,
    'Антонович (AED)': 796427.0,
}

DISPLAY_NAMES = {
    'TX7 (AED)': 'TX7 Solutions',
    'Фирсов (AED)': 'Фирсов',
    'D5 (AED)': 'D5',
    'K4 (AED)': 'K4',
    'Smart (AED)': 'Smart Renovations',
    'Антонович (AED)': 'Антонович',
}

COLORS = {
    'TX7 (AED)': '#4e89e8',
    'Фирсов (AED)': '#e84e7e',
    'D5 (AED)': '#e8a34e',
    'K4 (AED)': '#4ec9e8',
    'Smart (AED)': '#67c98c',
    'Антонович (AED)': '#b06cf4',
}

WEBSITES = {
    'TX7 (AED)': 'http://www.tx7solutions.com/',
    'Фирсов (AED)': 'https://dubai-firsov-design.com',
    'D5 (AED)': 'https://d5renovationdubai.com',
    'K4 (AED)': 'https://k4.ae',
    'Smart (AED)': 'https://smartrenovation.ae',
    'Антонович (AED)': 'https://antonovich-group.ae',
}

QUOTE_FILES = {
    'TX7 (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/Tareq/TX7_QT_45.docx'),
    'Фирсов (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/Фирсов/Смета по объекту.xlsx'),
    'D5 (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/D5/Villa 37 Golf Jumeirah Park- D5-Renovation Estimate.xlsx'),
    'K4 (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/К4/K4-5375_Sundails _ Villa 37 _ Jumeirah Golf Estates_K4-5375 (4).pdf'),
    'Smart (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/smart renovations Italian/Timur Vagizov - Jumeirah Golf Estates Villa Renovation EXECUTION.pdf'),
    'Антонович (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/Антонович/QUOTATION FOR INTERIOR FIT-OUT WORKS- DUBAI HILLS. 13.04.2026 copy.pdf'),
}

PORTFOLIO_DIRS = {
    'Фирсов (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/Фирсов/Портфолио фирсов'),
    'D5 (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/D5/Портфолио D5'),
    'Антонович (AED)': Path('/Users/Timur/Desktop/Общая папка с гугла/Khizar docs/Jumeirah golf estate/Renovation/КП подрядчиков/Антонович/Портфолио Антонович'),
}

FEEDBACK = {
    'TX7 (AED)': 'Кп выдал не выезжая на виллу, говорит что цена включает материалы местных поставщиков. С одной стороны он явный дебютант, с другой возможно он единственный не пытался угадать покупательную способность. На практике с ними скорее всего придется рядом стоять и не влезут в срок,',
    'Фирсов (AED)': 'у них странное КП, полы потеряли, напутали сами себя, включили в диагностику инженерки и ремонт сразу. Эти наверное самые гибкие, их можно прогнуть. У них есть кейс в Атлантис роял и большой дом на рублевке ремонтируют в данный момент. На Джумейра гольф у них была одна вилла, они доделывали ее за К4 (турками)',
    'D5 (AED)': 'самое подробное КП на текущий момент, вроде включено по-максимуму, портфолио обычное, русские',
    'K4 (AED)': 'это турки которые показывали нам виллу на пальме. Это очень большая компания, они могут много народа нагнать, но говорят что они сильно не вникают в процессы на площадке, слабая скорость реакции',
    'Smart (AED)': 'итальянец, КП почти ничего кроме демонтажей в себя не включает, с одной сторны инспекция и дизайн включены бесплатно, с другой ничего не понятно по ценам на работы которые они еще не брались оценивать и у них довольно много оговорок в договоре',
    'Антонович (AED)': 'это Туркмены и Узбеки, у них есть лакшери объекты в резюме, много вилл, КП неполное, его надо уточнять, но в целом цены не ломят',
}

# Подтверждено пользователем в этом проекте.
USER_CONFIRMED_CELL_NOTES = {
    ('Smart (AED)', 92): 'входит только демонтаж',
}

SECTION_ORDER = [16, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 17]
SECTION_TITLE = {
    16: '0. Мобилизация / управление проектом (ТЗ §16)',
    4: '1. Предремонтная диагностика (ТЗ §4)',
    5: '2. Демонтаж (ТЗ §5)',
    6: '3. Санузлы (ТЗ §6)',
    7: '4. Полы (ТЗ §7)',
    8: '5. Лестница (ТЗ §8)',
    9: '6. Двери (ТЗ §9)',
    10: '7. Кухня / хранение (ТЗ §10)',
    11: '8. Покраска (ТЗ §11)',
    12: '9. Бассейн (ТЗ §12)',
    13: '10. Кровля (ТЗ §13)',
    14: '11. Внешнее освещение (ТЗ §14)',
    17: '12. Контроль (ТЗ §17)',
}


def fmt_money(v: float | None) -> str:
    if v is None:
        return 'не могу подтвердить'
    s = f"{v:,.2f}" if abs(v - round(v)) > 1e-9 else f"{int(round(v)):,}"
    return s.replace(',', ' ')


def as_file_url(path: Path) -> str:
    return path.as_uri()


def eval_formula(val: str) -> float | None:
    t = val.strip().replace(' ', '')
    if not t.startswith('='):
        return None
    if not re.fullmatch(r'=[0-9+\-.]+', t):
        return None
    try:
        return float(eval(t[1:], {"__builtins__": {}}, {}))
    except Exception:
        return None


@dataclass
class CellView:
    cls: str
    text: str
    numeric: float | None


def normalize_cell(raw) -> CellView:
    if raw is None or raw == '':
        return CellView('v-empty', '—', None)

    if isinstance(raw, (int, float)):
        return CellView('v-num', fmt_money(float(raw)), float(raw))

    s = str(raw).strip()
    f = eval_formula(s)
    if f is not None:
        return CellView('v-num', fmt_money(f), f)

    low = s.lower()
    if 'не могу подтвердить' in low:
        return CellView('v-unc', 'не могу подтвердить', None)
    if 'не входит' in low:
        return CellView('v-out', 'не входит', None)
    if 'входит' in low or 'включ' in low or 'foc' in low:
        return CellView('v-ref', 'входит', None)
    if low == 'tba':
        return CellView('v-unc', 'TBA', None)
    return CellView('v-text', html.escape(s), None)


def section_key(raw) -> int | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = re.match(r'^(\d+)', s)
    if not m:
        return None
    return int(m.group(1))


def load_d5_total_from_quote(path: Path) -> float | None:
    if not path.exists():
        return None
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return None

    # Берем только прямой итог "TOTAL, AED excl. VAT 5%".
    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                if 'total' not in v.lower() or 'excl' not in v.lower():
                    continue
                for dc in range(c + 1, min(ws.max_column, c + 6) + 1):
                    num = ws.cell(r, dc).value
                    if isinstance(num, (int, float)):
                        return float(num)
    return None


def build() -> None:
    wb = load_workbook(EXCEL_PATH, data_only=False)
    ws = wb['Скоуп_для_ручного_ввода']
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    contractor_cols = []
    for c, h in enumerate(headers, start=1):
        if isinstance(h, str) and h.endswith('(AED)'):
            contractor_cols.append((h, c))

    rows = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if not isinstance(no, (int, float)):
            continue
        no = int(no)
        sec = section_key(ws.cell(r, 2).value)
        sub = ws.cell(r, 3).value
        item = ws.cell(r, 4).value
        if sec is None or not item:
            continue

        cells = {}
        for h, c in contractor_cols:
            cells[h] = normalize_cell(ws.cell(r, c).value)

        rows.append({
            'no': no,
            'section': sec,
            'sub': str(sub) if sub is not None else '',
            'item': str(item),
            'cells': cells,
        })

    section_buckets: dict[int, list[dict]] = defaultdict(list)
    for row in rows:
        section_buckets[row['section']].append(row)
    for sec in list(section_buckets.keys()):
        section_buckets[sec].sort(key=lambda x: x['no'])

    ordered_sections = [s for s in SECTION_ORDER if s in section_buckets]
    for s in sorted(section_buckets.keys()):
        if s not in ordered_sections:
            ordered_sections.append(s)

    comparable = {h: 0.0 for h, _ in contractor_cols}
    section_numeric: dict[str, dict[int, float]] = {h: defaultdict(float) for h, _ in contractor_cols}
    for row in rows:
        for h, _ in contractor_cols:
            n = row['cells'][h].numeric
            if n is not None:
                comparable[h] += n
                section_numeric[h][row['section']] += n

    quote_totals = dict(QUOTE_TOTALS)
    d5_col = 'D5 (AED)'
    d5_quote = QUOTE_FILES.get(d5_col)
    if d5_quote:
        d5_total = load_d5_total_from_quote(d5_quote)
        if d5_total is not None:
            quote_totals[d5_col] = d5_total

    demolition_only = {}
    for h, _ in contractor_cols:
        non_zero_sections = [sec for sec, v in section_numeric[h].items() if abs(v) > 1e-9]
        demolition_only[h] = len(non_zero_sections) == 1 and non_zero_sections[0] == 5

    # Top cards
    cards_html = []
    for h, _ in contractor_cols:
        q = quote_totals.get(h)
        comp = comparable[h]
        diff = q - comp if q is not None else None
        per_m2 = comp / VILLA_AREA_M2 if VILLA_AREA_M2 else None
        if diff is not None and abs(diff) < 0.01:
            flag_cls = 'ok'
            flag_text = 'Материалы/опции не выделены'
        elif diff is not None and diff > 0:
            flag_cls = 'warn'
            flag_text = 'Материалы/опции выделены'
        else:
            flag_cls = 'warn'
            flag_text = 'не могу подтвердить'

        cards_html.append(f'''
    <div class="card" style="--c:{COLORS.get(h, '#7785a5')}">
      <div class="name">{html.escape(DISPLAY_NAMES.get(h, h))}</div>
      <div class="line comparable"><span>Сравнимая сумма</span><b>{fmt_money(comp)} AED</b></div>
      <div class="line"><span>Полная КП</span><b>{fmt_money(q)} AED</b></div>
      <div class="line"><span>Материалы/опции</span><b>{fmt_money(diff)} AED</b></div>
      <div class="line"><span>Работы за м²</span><b class="per-m2">{fmt_money(per_m2)} AED/м²</b></div>
      <div class="flag {flag_cls}">{flag_text}</div>
      {('<div class="flag only-demo">входит только демонтаж</div>' if demolition_only.get(h) else '')}
    </div>''')

    # Table
    col_count = 1 + len(contractor_cols)
    company_headers = ''.join(f'<th>{html.escape(DISPLAY_NAMES.get(h, h))}</th>' for h, _ in contractor_cols)
    first_col = 30.0
    other_col = (100.0 - first_col) / len(contractor_cols)

    body_parts = []
    for sec in ordered_sections:
        sec_id = f's{sec}'
        title = SECTION_TITLE.get(sec, f'Раздел {sec}')
        body_parts.append(
            f'<tr class="section-head" onclick="toggleSection(\'{sec_id}\')"><td colspan="{col_count}">{html.escape(title)} <span class="toggle" id="{sec_id}-icon">▼</span></td></tr>'
        )

        sec_totals = {h: 0.0 for h, _ in contractor_cols}
        for row in section_buckets[sec]:
            for h, _ in contractor_cols:
                n = row['cells'][h].numeric
                if n is not None:
                    sec_totals[h] += n

        tds = ''.join(f'<td class="v-sec-total">{fmt_money(sec_totals[h]) if abs(sec_totals[h])>1e-9 else "—"}</td>' for h, _ in contractor_cols)
        body_parts.append(f'<tr class="item-row section-total {sec_id}"><td class="sticky">ИТОГО ПО РАЗДЕЛУ</td>{tds}</tr>')

        for row in section_buckets[sec]:
            label = f'<span class="n">{row["no"]}</span> {html.escape(row["item"])}<div class="sub">ТЗ {html.escape(str(row["section"]))}.{html.escape(str(row["sub"]))}</div>'
            cell_parts = []
            for h, _ in contractor_cols:
                cell = row['cells'][h]
                note = USER_CONFIRMED_CELL_NOTES.get((h, row['no']))
                note_html = f'<div class="cell-note">{html.escape(note)}</div>' if note else ''
                cell_parts.append(f'<td class="{cell.cls}">{cell.text}{note_html}</td>')
            cells_html = ''.join(cell_parts)
            body_parts.append(f'<tr class="item-row {sec_id}"><td class="sticky">{label}</td>{cells_html}</tr>')

    grand = ''.join(f'<td>{fmt_money(comparable[h])}</td>' for h, _ in contractor_cols)
    body_parts.append(f'<tr class="grand-total"><td>ИТОГО ПО ТАБЛИЦЕ РАЗДЕЛОВ (числовые значения)</td>{grand}</tr>')

    # Sources section
    source_rows = []
    for h, _ in contractor_cols:
        site = WEBSITES.get(h)
        site_html = f'<a href="{html.escape(site)}" target="_blank" rel="noopener">{html.escape(site)}</a>' if site else 'не могу подтвердить'

        kp_link = DROPBOX_KP_LINKS.get(h)
        if kp_link:
            kp_html = f'<a href="{html.escape(kp_link)}" target="_blank" rel="noopener">открыть в Dropbox</a>'
        else:
            kp_html = 'не могу подтвердить'

        portfolio_link = DROPBOX_PORTFOLIO_LINKS.get(h)
        if portfolio_link:
            local_html = f'<a href="{html.escape(portfolio_link)}" target="_blank" rel="noopener">открыть в Dropbox</a>'
        else:
            local_html = 'не могу подтвердить'

        demo_mark = 'входит только демонтаж' if demolition_only.get(h) else '—'
        source_rows.append(
            f'<tr><td><b>{html.escape(DISPLAY_NAMES.get(h, h))}</b></td><td>{site_html}</td><td>{kp_html}</td><td>{local_html}</td><td>{demo_mark}</td></tr>'
        )

    feedback_rows = []
    for h, _ in contractor_cols:
        fb = FEEDBACK.get(h, 'не могу подтвердить')
        feedback_rows.append(f'<tr><td><b>{html.escape(DISPLAY_NAMES.get(h,h))}</b></td><td class="rec-cell">{html.escape(fb)}</td></tr>')

    html_out = f'''<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1.0" />
<title>Сравнение КП подрядчиков — по разделам ТЗ</title>
<style>
* {{ box-sizing: border-box; }}
body {{ margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",system-ui,sans-serif; background:#f1f3f7; color:#1a2233; }}
.header {{ position:sticky; top:0; z-index:50; background:#141a2b; color:#fff; padding:14px 18px; box-shadow:0 2px 8px rgba(0,0,0,.25); }}
.header-row {{ display:flex; align-items:center; justify-content:space-between; gap:12px; }}
.header h1 {{ margin:0; font-size:18px; }}
.header p {{ margin:4px 0 0; font-size:12px; color:#aab4d6; }}
.pdf-btn {{ display:inline-block; border:1px solid #3a4a78; background:#1e2741; color:#fff; text-decoration:none; border-radius:8px; padding:8px 10px; font-size:12px; font-weight:700; white-space:nowrap; }}
.pdf-btn:hover {{ background:#253054; }}
.container {{ max-width:1780px; margin:0 auto; padding:16px; }}
.cards {{ display:grid; grid-template-columns:repeat(3,minmax(240px,1fr)); gap:10px; margin-bottom:14px; }}
@media (max-width:1200px) {{ .cards {{ grid-template-columns:repeat(2,minmax(240px,1fr)); }} }}
.card {{ background:#fff; border-radius:12px; border-top:4px solid var(--c); padding:8px 10px; box-shadow:0 2px 8px rgba(0,0,0,.08); }}
.card .name {{ font-weight:800; font-size:15px; margin-bottom:6px; }}
.card .line {{ display:flex; justify-content:space-between; gap:8px; margin:2px 0; font-size:12px; align-items:flex-start; }}
.card .line.comparable {{ background:#ecf8ef; border:1px solid #c8ebd8; border-radius:8px; padding:4px 6px; margin:4px 0; }}
.card .line span {{ color:#5e6b85; line-height:1.2; }}
.card .line b {{ color:#10182d; text-align:right; overflow-wrap:anywhere; font-size:16px; line-height:1.1; }}
.card .line b.per-m2 {{ font-size:12px; font-weight:700; }}
.flag {{ margin-top:6px; padding:5px 7px; border-radius:8px; font-size:10px; font-weight:700; }}
.flag.warn {{ background:#fff4d8; color:#7a5800; }}
.flag.ok {{ background:#e6f8ee; color:#1d6b45; }}
.flag.only-demo {{ background:#fde7e7; color:#9b1f1f; }}

.legend {{ background:#fff; border-radius:10px; padding:10px 12px; margin-bottom:12px; box-shadow:0 2px 8px rgba(0,0,0,.06); display:flex; flex-wrap:wrap; gap:12px; font-size:12px; }}
.legend .it {{ display:flex; align-items:center; gap:6px; }}
.legend .sw {{ width:14px; height:14px; border-radius:3px; border:1px solid #ccd3e0; }}

.table-wrap {{ background:#fff; border-radius:12px; box-shadow:0 2px 8px rgba(0,0,0,.08); overflow-x:auto; -webkit-overflow-scrolling:touch; }}
table {{ width:100%; min-width:980px; border-collapse:collapse; table-layout:fixed; }}
#mainTable thead th {{ background:#171d31; color:#fff; padding:10px 8px; font-size:12px; border-bottom:3px solid #303a5a; }}
#mainTable thead th:nth-child(1) {{ width:{first_col:.4f}%; text-align:left; }}
#mainTable thead th:nth-child(n+2) {{ width:{other_col:.4f}%; }}
tr.section-head td {{ background:#dfe3ef; padding:8px 10px; font-weight:800; font-size:13px; border-top:2px solid #c9d1e3; cursor:pointer; }}
tr.section-head .toggle {{ float:right; color:#4d5878; }}
tr.item-row td {{ border-bottom:1px solid #edf1f7; padding:7px 8px; font-size:12px; text-align:center; vertical-align:top; }}
tr.item-row td.sticky {{ text-align:left; overflow-wrap:anywhere; word-break:break-word; }}
tr.item-row .n {{ display:inline-block; min-width:30px; color:#7a859d; font-weight:700; }}
tr.item-row .sub {{ margin-top:2px; font-size:10px; color:#97a1b8; }}
tr.section-total td {{ font-weight:800; border-bottom:2px solid #dbe1ee; }}
.v-sec-total {{ color:#1b2438; font-weight:800; }}
.v-num {{ color:#1b2438; font-weight:700; background:transparent; }}
.v-ref {{ color:#1d6d44; font-weight:700; background:transparent; }}
.v-out {{ color:#992121; font-weight:700; background:transparent; }}
.v-unc {{ color:#0f558b; font-weight:700; background:transparent; }}
.v-text {{ color:#4a5675; font-weight:700; background:transparent; }}
.v-empty {{ color:#a2a8b8; background:transparent; }}
.cell-note {{ margin-top:3px; font-size:10px; line-height:1.2; color:#9b1f1f; font-weight:800; }}
tr.grand-total td {{ background:#171d31; color:#fff; font-weight:800; border-top:3px solid #0e1220; padding:10px 8px; }}
tr.grand-total td:first-child {{ text-align:left; }}

.summary {{ margin-top:14px; background:#fff; border-radius:12px; box-shadow:0 2px 8px rgba(0,0,0,.08); padding:12px; }}
.summary h2 {{ margin:0 0 8px; font-size:14px; }}
.summary th, .summary td {{ border-bottom:1px solid #edf1f7; padding:7px 8px; font-size:12px; text-align:left; vertical-align:top; }}
.summary th {{ background:#f3f6fc; color:#1b2438; }}
.rec-cell {{ line-height:1.45; }}
@media (max-width:900px) {{
  .header {{ position:static; padding:10px 12px; }}
  .header-row {{ align-items:flex-start; flex-direction:column; }}
  .header h1 {{ font-size:16px; }}
  .header p {{ font-size:11px; }}
  .pdf-btn {{ padding:7px 9px; font-size:11px; }}
  .container {{ padding:10px; }}
  .cards {{ grid-template-columns:1fr; gap:8px; }}
  .card {{ padding:8px; }}
  .card .name {{ font-size:14px; }}
  .card .line b {{ font-size:15px; }}
  .legend {{ font-size:11px; padding:8px 10px; }}
  #mainTable thead th {{ font-size:11px; padding:8px 6px; }}
  tr.item-row td {{ font-size:11px; padding:6px 6px; }}
  tr.item-row .sub {{ font-size:9px; }}
  .summary {{ padding:8px; }}
  .summary th, .summary td {{ font-size:11px; padding:6px; }}
}}
@media print {{
  .header {{ position:static; box-shadow:none; }}
  .pdf-btn {{ display:none; }}
  a, a:visited {{ color:#0a58ca; text-decoration:underline; }}
  .summary a::after {{ content:" (" attr(href) ")"; color:#4b556d; font-size:10px; word-break:break-all; }}
}}
</style>
</head>
<body>
<div class="header">
  <div class="header-row">
    <h1>Сравнение КП подрядчиков — по разделам ТЗ</h1>
    <a class="pdf-btn" href="{OUTPUT_PDF_NAME}" download>Скачать PDF</a>
  </div>
  <p>Источник строк: Scope_Manual_Input_from_TZ.xlsx. Площадь для расчета стоимости работ: {int(VILLA_AREA_M2)} м².</p>
</div>
<div class="container">
  <div class="cards">{''.join(cards_html)}
  </div>

  <div class="legend">
    <div class="it"><span class="sw" style="background:#ecf8ef"></span>Входит</div>
    <div class="it"><span class="sw" style="background:#fdeeee"></span>Не входит</div>
    <div class="it"><span class="sw" style="background:#e8f3ff"></span>Не могу подтвердить</div>
  </div>

  <div class="table-wrap">
    <table id="mainTable">
      <thead>
        <tr>
          <th>Позиция ТЗ</th>
          {company_headers}
        </tr>
      </thead>
      <tbody>
        {''.join(body_parts)}
      </tbody>
    </table>
  </div>

  <div class="summary">
    <h2>Сайты, КП и портфолио</h2>
    <table>
      <thead><tr><th>Подрядчик</th><th>Сайт</th><th>КП</th><th>Портфолио</th><th>Пометка</th></tr></thead>
      <tbody>
        {''.join(source_rows)}
      </tbody>
    </table>
  </div>

  <div class="summary">
    <h2>Фидбек заказчика</h2>
    <table>
      <thead><tr><th>Подрядчик</th><th>Комментарий</th></tr></thead>
      <tbody>
        {''.join(feedback_rows)}
      </tbody>
    </table>
  </div>
</div>
<script>
function toggleSection(id) {{
  const rows=document.querySelectorAll('.'+id);
  const icon=document.getElementById(id+'-icon');
  let hidden=false;
  rows.forEach(r=>{{
    r.style.display=(r.style.display==='none')?'':'none';
    hidden=r.style.display==='none';
  }});
  if(icon) icon.textContent=hidden?'▶':'▼';
}}
</script>
</body>
</html>
'''

    OUTPUT_HTML.write_text(html_out, encoding='utf-8')
    print(f'written: {OUTPUT_HTML}')


if __name__ == '__main__':
    build()
